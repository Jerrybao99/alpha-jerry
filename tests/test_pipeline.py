"""Step 1.3 单测——cache / format / pipeline，全 mock 不触网络（FR-DATA-05/06）。

覆盖：缓存键/读写/禁用、百分比格式化与输出行、采集编排（成功+缓存命中+失败隔离）、
失败清单落 xlsx、字段标准化顺序。
"""

from __future__ import annotations

import datetime as _dt
from concurrent.futures import Future, ThreadPoolExecutor
from pathlib import Path

from src.config import Settings
from src.core.pipeline import CollectionPipeline
from src.schemas.financial import ALL_OUTPUT_COLUMNS, PERCENT_FIELDS, StockFeatures, StockInfo
from src.utils.cache import Cache
from src.utils.format import format_percent, to_output_row


# ===== 测试桩 =====
class _SeqExecutor:
    """顺序执行器：submit 立即同步执行，便于确定性单测。"""

    def submit(self, fn, *args):
        f: Future = Future()
        f.set_running_or_notify_cancel()
        try:
            f.set_result(fn(*args))
        except Exception as exc:  # noqa: BLE001
            f.set_exception(exc)
        return f

    def shutdown(self, wait=False) -> None:  # noqa: ANN001
        pass


class _FakeFetcher:
    """可控 BaseFetcher：按 ts_code 返回脚本化特征或抛异常。"""

    def __init__(self) -> None:
        self.features: dict[str, StockFeatures] = {}
        self.raise_on: set[str] = set()
        self.none_on: set[str] = set()
        self.calls: list[str] = []

    def fetch_stock_list(self) -> list[StockInfo]:
        codes = set(self.features) | self.raise_on | self.none_on
        return [StockInfo(ts_code=c, symbol=c.split(".")[0], name=c) for c in sorted(codes)]

    def fetch_financials(self, ts_code, period=None):
        self.calls.append(ts_code)
        if ts_code in self.raise_on:
            raise RuntimeError(f"boom {ts_code}")
        if ts_code in self.none_on:
            return None
        f = self.features[ts_code]
        return StockFeatures(**f.model_dump())


def _feat(ts_code: str, **kw) -> StockFeatures:
    return StockFeatures(ts_code=ts_code, **kw)


def _settings(data_dir: Path) -> Settings:
    return Settings(tushare_token="t", data_dir=str(data_dir), concurrency=2)


# ===== Cache =====
def test_cache_key_for() -> None:
    assert Cache.key_for("600000.SH", "20241231") == "600000.SH_20241231"
    assert Cache.key_for("600000.SH", None) == "600000.SH_latest"


def test_cache_set_get_roundtrip(tmp_path: Path) -> None:
    cache = Cache(tmp_path / "c")
    feat = _feat("600000.SH", revenue=1.5e10, roe=12.5)
    cache.set("600000.SH", "20241231", feat)
    got = cache.get("600000.SH", "20241231")
    assert got is not None
    assert got.revenue == 1.5e10
    assert got.roe == 12.5


def test_cache_miss_returns_none(tmp_path: Path) -> None:
    assert Cache(tmp_path / "c").get("000001.SZ", None) is None


def test_cache_disabled(tmp_path: Path) -> None:
    cache = Cache(tmp_path / "c", enabled=False)
    cache.set("600000.SH", None, _feat("600000.SH"))
    assert cache.get("600000.SH", None) is None
    assert not (tmp_path / "c").exists() or not list((tmp_path / "c").iterdir())


def test_cache_corrupt_returns_none(tmp_path: Path) -> None:
    cache = Cache(tmp_path / "c")
    (tmp_path / "c" / "600000.SH_latest.json").write_text("not json", encoding="utf-8")
    assert cache.get("600000.SH", None) is None


# ===== format =====
def test_format_percent() -> None:
    assert format_percent(30.5) == "30.50%"
    assert format_percent(0) == "0.00%"
    assert format_percent(None) is None


def test_to_output_row_columns_order_and_percent() -> None:
    feat = _feat("600000.SH", netprofit_yoy=20.0, roe=12.5, revenue=1.0e10)
    row = to_output_row(feat)
    assert list(row.keys()) == list(ALL_OUTPUT_COLUMNS)
    assert row["ts_code"] == "600000.SH"
    assert row["netprofit_yoy"] == "20.00%"  # 百分比字段格式化
    assert row["roe"] == 12.5  # 非百分比保持原值
    assert row["revenue"] == 1.0e10
    assert row["eps"] is None


def test_to_output_row_all_percent_fields_formatted() -> None:
    feat = _feat("600000.SH", **{f: 1.0 for f in PERCENT_FIELDS})
    row = to_output_row(feat)
    for f in PERCENT_FIELDS:
        assert row[f] == "1.00%", f


# ===== pipeline =====
def test_pipeline_collects_successes(tmp_path: Path) -> None:
    fetcher = _FakeFetcher()
    fetcher.features = {
        "600000.SH": _feat("600000.SH", revenue=1.0e10),
        "000001.SZ": _feat("000001.SZ", revenue=2.0e10),
    }
    pipe = CollectionPipeline(
        fetcher, settings=_settings(tmp_path), cache=Cache(tmp_path / "c"), executor=_SeqExecutor()
    )
    result = pipe.run()
    assert result.total == 2
    assert result.success_count == 2
    assert result.failure_count == 0
    assert result.cached_hits == 0
    assert [s.ts_code for s in result.successes] == ["000001.SZ", "600000.SH"]


def test_pipeline_cache_hit_on_second_run(tmp_path: Path) -> None:
    fetcher = _FakeFetcher()
    fetcher.features = {"600000.SH": _feat("600000.SH", revenue=1.0e10)}
    cache = Cache(tmp_path / "c")
    pipe = CollectionPipeline(fetcher, settings=_settings(tmp_path), cache=cache, executor=_SeqExecutor())
    pipe.run()
    assert fetcher.calls == ["600000.SH"]
    # 第二次跑：命中缓存，不再调接口
    result2 = pipe.run()
    assert result2.cached_hits == 1
    assert fetcher.calls == ["600000.SH"]  # 未重复调接口


def test_pipeline_failure_isolation(tmp_path: Path) -> None:
    """单股抛异常不拖垮整体，记入失败清单。"""
    fetcher = _FakeFetcher()
    fetcher.features = {
        "600000.SH": _feat("600000.SH"),
        "000001.SZ": _feat("000001.SZ"),
    }
    fetcher.raise_on = {"000001.SZ"}
    pipe = CollectionPipeline(
        fetcher, settings=_settings(tmp_path), cache=Cache(tmp_path / "c"), executor=_SeqExecutor()
    )
    result = pipe.run()
    assert result.success_count == 1
    assert result.failure_count == 1
    assert result.failures[0].ts_code == "000001.SZ"
    assert "boom" in result.failures[0].error


def test_pipeline_none_features_recorded_as_failure(tmp_path: Path) -> None:
    fetcher = _FakeFetcher()
    fetcher.features = {"600000.SH": _feat("600000.SH")}
    fetcher.none_on = {"600000.SH"}
    pipe = CollectionPipeline(
        fetcher, settings=_settings(tmp_path), cache=Cache(tmp_path / "c"), executor=_SeqExecutor()
    )
    result = pipe.run()
    assert result.success_count == 0
    assert result.failures[0].error == "无数据"


def test_pipeline_codes_filter(tmp_path: Path) -> None:
    """codes 非空时只采指定股票（冒烟测试场景）。"""
    fetcher = _FakeFetcher()
    fetcher.features = {c: _feat(c) for c in ("600000.SH", "000001.SZ", "000002.SZ")}
    pipe = CollectionPipeline(
        fetcher, settings=_settings(tmp_path), cache=Cache(tmp_path / "c"), executor=_SeqExecutor()
    )
    result = pipe.run(codes=["600000.SH"])
    assert result.total == 1
    assert result.successes[0].ts_code == "600000.SH"


def test_pipeline_to_rows(tmp_path: Path) -> None:
    fetcher = _FakeFetcher()
    fetcher.features = {"600000.SH": _feat("600000.SH", netprofit_yoy=20.0)}
    pipe = CollectionPipeline(
        fetcher, settings=_settings(tmp_path), cache=Cache(tmp_path / "c"), executor=_SeqExecutor()
    )
    result = pipe.run()
    rows = pipe.to_rows(result)
    assert len(rows) == 1
    assert list(rows[0].keys()) == list(ALL_OUTPUT_COLUMNS)
    assert rows[0]["netprofit_yoy"] == "20.00%"


def test_pipeline_write_failures(tmp_path: Path) -> None:
    fetcher = _FakeFetcher()
    fetcher.features = {"600000.SH": _feat("600000.SH")}
    fetcher.raise_on = {"000001.SZ"}
    pipe = CollectionPipeline(
        fetcher, settings=_settings(tmp_path), cache=Cache(tmp_path / "c"), executor=_SeqExecutor()
    )
    result = pipe.run()
    out = pipe.write_failures(result.failures, date=_dt.date(2026, 7, 23))
    assert out.name == "260723-失败.xlsx"
    assert out.exists()
    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("ts_code", "name", "error")
    assert rows[1][0] == "000001.SZ"
    assert "boom" in rows[1][2]


def test_pipeline_real_threadpool_smoke(tmp_path: Path) -> None:
    """真实 ThreadPoolExecutor（concurrency=2）下也能正确聚合。"""
    fetcher = _FakeFetcher()
    fetcher.features = {f"60000{i}.SH": _feat(f"60000{i}.SH") for i in range(4)}
    exe = ThreadPoolExecutor(max_workers=2)
    pipe = CollectionPipeline(fetcher, settings=_settings(tmp_path), cache=Cache(tmp_path / "c"), executor=exe)
    try:
        result = pipe.run()
    finally:
        exe.shutdown()
    assert result.success_count == 4
    assert result.failure_count == 0
